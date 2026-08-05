import pandas as pd
import glob
import os
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# ==========================================
#        CONFIGURACIÓN MANUAL
# ==========================================
ruta_carpeta = r'C:\Users\mendozapa\HITSS\Angel Jesus Zavala Ubillus - Fotos Compartido\Cortes Agosto'
ruta_logo = r'LogoHits.png'

usuarios_permitidos = [
    'E759763',  # ERICK JONATHAN NÚÑEZ LUDEÑA
    'E759708',  # FRESIA CLEMENTE RODRIGUEZ
    'E759762',  # RENATO PAULO CALLAN ANDRADE
    'E759747',  # JORGELUIS ARMANDO CORDOVA TORRES
    #'E759761',  # PABLO ANDRÉS MENDOZA CALLE
    'E759899',  # AARON ADRIANO GUZMAN PHATTE
    'E760568',  # RICHARD BRUSS HUAYSARA ROMAN
    'E760642',  # MANUEL ALEJANDRO ANGELES RAMON
    'E760214',  # Renzo Alfredo Ulloa Bozeta 
]

# --- CÁLCULO AUTOMÁTICO DE FECHA Y HORA DE CORTE ---
ahora = datetime.now()
hora_corte_manual = ahora.strftime('%H:00')  
fecha_hoy_str = ahora.strftime('%d/%m/%Y')   

print(f"[INFO] Fecha actual detectada: {fecha_hoy_str} -> Hora de corte aplicada: {hora_corte_manual}")

ruta_salida = os.path.join(ruta_carpeta, 'Reporte_Consolidado_Final.xlsx')
# ==========================================


# --- PARTE 0: AUTO-REUBICACIÓN DESDE DESCARGAS (SOPORTA MÚLTIPLES ARCHIVOS) ---
ruta_descargas = os.path.join(os.path.expanduser('~'), 'Downloads')
archivos_descargados = glob.glob(os.path.join(ruta_descargas, "*REGISTRO DE RF FOTOS*.xlsx"))

if archivos_descargados:
    for arch in archivos_descargados:
        try:
            nombre_base = os.path.basename(arch)
            shutil.move(arch, os.path.join(ruta_carpeta, nombre_base))
            print(f"[ÉXITO] Archivo trasladado desde Descargas: {nombre_base}")
        except Exception as e:
            print(f"[ERROR] No se pudo mover {arch}: {e}")
else:
    print("[INFO] No hay nuevas descargas. Se usará(n) el/los 'REGISTRO DE RF FOTOS' existentes en la carpeta.")


# --- PARTE 1: PROCESAR "REPORTE NIVELES IA" ---
archivos_cortes = glob.glob(os.path.join(ruta_carpeta, "Reporte_Niveles_IA*.xlsx"))
datos_reporte = []
meses_espanol = {1:'ene', 2:'feb', 3:'mar', 4:'abr', 5:'may', 6:'jun', 
                 7:'jul', 8:'ago', 9:'sep', 10:'oct', 11:'nov', 12:'dic'}

for archivo in archivos_cortes:
    try:
        print(f"[LECTURA] Procesando reporte IA: {os.path.basename(archivo)}")
        
        # Extrae la fecha ignorando el posible guion bajo inicial
        match = re.search(r'Reporte_Niveles_IA_?(.*?)\.xlsx', os.path.basename(archivo), re.IGNORECASE)
        fecha_texto = match.group(1).strip() if match else "Desconocida"
        
        try:
            # Convierte el formato DDMMYYYY (ej. 05082026) a fecha
            f_obj = pd.to_datetime(fecha_texto, format='%d%m%Y')
            fecha = f"{f_obj.day:02d}-{meses_espanol[f_obj.month]}"
        except Exception:
            fecha = fecha_texto
        
        # Identificar y leer la hoja "Corte x asesor" independientemente de mayúsculas/minúsculas
        xls = pd.ExcelFile(archivo)
        hoja_objetivo = next((s for s in xls.sheet_names if 'corte x asesor' in s.strip().lower()), None)
        
        if hoja_objetivo:
            df_control = pd.read_excel(archivo, sheet_name=hoja_objetivo)
        else:
            print(f"[AVISO] No se encontró la hoja 'Corte x asesor' en {os.path.basename(archivo)}. Usando la primera hoja.")
            df_control = pd.read_excel(archivo)
            
        # Búsqueda dinámica de la fila de cabeceras que contenga 'TOTAL' y 'PENDIENTES'
        header_idx = -1
        for i, row in df_control.head(20).iterrows():
            row_str = ' '.join([str(val).upper() for val in row if pd.notna(val)])
            if 'TOTAL' in row_str and 'PENDIENTE' in row_str:
                header_idx = i
                break
                
        if header_idx != -1:
            # Asignar la fila encontrada como nombres de columnas
            df_control.columns = df_control.iloc[header_idx]
            # Recortar el DataFrame para que solo contenga los datos debajo de las cabeceras
            df_control = df_control.iloc[header_idx+1:].reset_index(drop=True)
        else:
            print(f"[AVISO] No se encontraron cabeceras claras de TOTAL y PENDIENTES en {os.path.basename(archivo)}")

        # Identificar las columnas exactas
        col_total = next((c for c in df_control.columns if pd.notna(c) and 'TOTAL' in str(c).strip().upper()), None)
        col_pend = next((c for c in df_control.columns if pd.notna(c) and 'PENDIENTE' in str(c).strip().upper()), None)
        col_term = next((c for c in df_control.columns if pd.notna(c) and 'TERMINADO' in str(c).strip().upper()), None)
        
        # Extraer el primer valor numérico válido de cada columna
        try:
            total = pd.to_numeric(df_control[col_total], errors='coerce').dropna().iloc[0] if col_total else 0
        except IndexError:
            total = 0
            
        try:
            pendientes = pd.to_numeric(df_control[col_pend], errors='coerce').dropna().iloc[0] if col_pend else 0
        except IndexError:
            pendientes = 0
            
        if col_term:
            try:
                terminados = pd.to_numeric(df_control[col_term], errors='coerce').dropna().iloc[0]
            except IndexError:
                terminados = total - pendientes
        else:
            terminados = total - pendientes
            
        if int(total) == 0 and int(pendientes) == 0:
            print(f"[OMITIDO] Sin datos encontrados para procesar en: {os.path.basename(archivo)}")
            continue
            
        datos_reporte.append({
            'FECHA': fecha, 
            'TOTAL': int(total), 
            'PENDIENTES': int(pendientes),
            'TERMINADOS': int(terminados)
        })
    except Exception as e:
        print(f"Error procesando {archivo}: {e}")

df_cortes = pd.DataFrame(datos_reporte)
if not df_cortes.empty:
    column_order = ['FECHA', 'TOTAL', 'PENDIENTES', 'TERMINADOS']
    df_cortes = df_cortes[column_order]
    
    totales = df_cortes[['TOTAL', 'PENDIENTES', 'TERMINADOS']].sum()
    df_cortes = pd.concat([df_cortes, pd.DataFrame({'FECHA': ['TOTAL'], **totales.to_dict()})], ignore_index=True)


# --- PARTE 2: PROCESAR "REGISTRO DE RF FOTOS" (COMBINA TODOS LOS ARCHIVOS QUE COINCIDAN) ---
archivos_fotos = glob.glob(os.path.join(ruta_carpeta, "*REGISTRO DE RF FOTOS*.xlsx"))
df_fotos_dinamica = pd.DataFrame()

if archivos_fotos:
    try:
        lista_dfs = []
        for arch in archivos_fotos:
            print(f"[LECTURA] Leyendo base de fotos: {os.path.basename(arch)}")
            df_temp = pd.read_excel(arch)
            df_temp.columns = df_temp.columns.astype(str).str.strip()
            lista_dfs.append(df_temp)
        
        # Consolidación de todas las partes cargadas
        df_f = pd.concat(lista_dfs, ignore_index=True)
        
        col_fecha = next((col for col in df_f.columns if 'hora de fi' in col.lower()), None)
        col_hora = next((col for col in df_f.columns if 'nalización' in col.lower()), None)
        
        if 'HORA DE FINALIZACIÓN' in df_f.columns:
            col_fecha = col_hora = 'HORA DE FINALIZACIÓN'
            
        if col_fecha and col_hora:
            if 'USUARIO E' in df_f.columns and usuarios_permitidos:
                df_f['USUARIO E'] = df_f['USUARIO E'].astype(str).str.strip()
                df_f = df_f[df_f['USUARIO E'].isin(usuarios_permitidos)]
            
            fechas_evaluadas = pd.to_datetime(df_f[col_fecha], errors='coerce').dt.strftime('%d/%m/%Y')
            df_f = df_f[fechas_evaluadas == fecha_hoy_str]
            
            horas_str = df_f[col_hora].astype(str).str.replace('a. m.', 'AM', regex=False).str.replace('p. m.', 'PM', regex=False)
            horas_24h = pd.to_datetime(horas_str, errors='coerce').dt.strftime('%H:%M:%S')
            limite_str = pd.to_datetime(hora_corte_manual, format='%H:%M').strftime('%H:%M:%S')
            
            df_f = df_f[horas_24h.fillna('23:59:59') < limite_str]
            
            if not df_f.empty:
                df_f['HORA_INT'] = pd.to_datetime(horas_str, errors='coerce').dt.hour
                df_f['FECHA_SOLO'] = f"- {fecha_hoy_str}"
                
                df_grouped = df_f.groupby(['FECHA_SOLO', 'USUARIO E', 'HORA_INT']).size().reset_index(name='SOT')
                df_ajustado_list = []
                
                for (fecha_aux, usuario_aux), group in df_grouped.groupby(['FECHA_SOLO', 'USUARIO E']):
                    conteo = dict(zip(group['HORA_INT'], group['SOT']))
                    for h_baja in [h for h, c in conteo.items() if 1 <= c <= 5]:
                        if conteo[h_baja] == 0:
                            continue
                        objetivos = [h for h, c in conteo.items() if h != h_baja and c > 0]
                        if objetivos:
                            conteo[min(objetivos, key=lambda x: abs(x - h_baja))] += conteo[h_baja]
                        conteo[h_baja] = 0
                        
                    for h, c in conteo.items():
                        if c > 5:
                            df_ajustado_list.append({'FECHA_SOLO': fecha_aux, 'USUARIO E': usuario_aux, 'HORA_INT': h, 'SOT': c})
                
                if df_ajustado_list:
                    df_ajustada = pd.DataFrame(df_ajustado_list)
                    
                    def mapear_hora_string(h_int):
                        am_pm = 'a. m.' if h_int < 12 else 'p. m.'
                        h_disp = h_int if h_int <= 12 else h_int - 12
                        return f"{h_disp if h_disp != 0 else 12:02d}:00 {am_pm}"
                    
                    df_ajustada['HORA_FORMATO'] = df_ajustada['HORA_INT'].apply(mapear_hora_string)
                    horas_ordenadas = [mapear_hora_string(h) for h in sorted(df_ajustada['HORA_INT'].unique())]
                    
                    df_ajustada['HORA_FORMATO'] = pd.Categorical(df_ajustada['HORA_FORMATO'], categories=horas_ordenadas, ordered=True)
                    df_fotos_dinamica = pd.pivot_table(df_ajustada, index=['USUARIO E'], columns='HORA_FORMATO', values='SOT', aggfunc='sum', margins=True, margins_name='Total').fillna("")
            else:
                print(f"[AVISO] Después de filtrar por fecha de hoy ({fecha_hoy_str}) y hora (<{hora_corte_manual}), no quedaron registros en RF FOTOS.")
        else:
            print(f"No se detectaron las columnas requeridas. Halladas: {list(df_f.columns)}")
    except Exception as e:
        print(f"Error procesando fotos: {e}")


# --- PARTE 3: EXPORTAR Y APLICAR FORMATO CONJUNTO ---
with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
    startrow_cortes = 2 
    if not df_cortes.empty:
        df_cortes.to_excel(writer, sheet_name='Resumen', index=False, startrow=startrow_cortes)
        fila_t2 = startrow_cortes + len(df_cortes) + 3
    else:
        fila_t2 = startrow_cortes
        
    if not df_fotos_dinamica.empty:
        df_fotos_dinamica.to_excel(writer, sheet_name='Resumen', index=True, startrow=fila_t2)


# --- PARTE 4: MEJORAS VISUALES ---
if os.path.exists(ruta_salida):
    wb = load_workbook(ruta_salida)
    ws = wb['Resumen']
    
    # --- Paleta de Colores ---
    banner_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul Oscuro
    blue_fill   = PatternFill(start_color="3B5E94", end_color="3B5E94", fill_type="solid") # Azul Base Tablas
    pend_fill   = PatternFill(start_color="558ED5", end_color="558ED5", fill_type="solid") # Azul Claro Pendientes
    
    white_font  = Font(color="FFFFFF", bold=True)
    banner_font = Font(color="FFFFFF", bold=True, size=16) 
    
    center_align = Alignment(horizontal="center", vertical="center")

    # 1. Alineación general centrada para todas las celdas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align

    # 2. INSERTAR ENCABEZADO BANNER Y LOGO
    max_col = ws.max_column if ws.max_column > 1 else 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    
    for col_idx in range(1, max_col + 1):
        celda = ws.cell(row=1, column=col_idx)
        celda.fill = banner_fill
        celda.font = banner_font
            
    # Texto centrado limpiamente
    ws.cell(row=1, column=1).value = f"CORTE FOTOS - {hora_corte_manual}"
    
    # Altura del Banner ajustada a 50 puntos
    ws.row_dimensions[1].height = 50

    # Insertar Logo en la Celda A1
    if os.path.exists(ruta_logo):
        try:
            img = Image(ruta_logo)
            img.height = 64
            img.width = 64
            img.left = 80 
            img.top = 3   
            
            ws.add_image(img, 'A1')
        except Exception as e:
            pass

    # 3. Formato Tabla 1 (Reporte Niveles IA)
    if not df_cortes.empty:
        row_header_cortes = startrow_cortes + 1                  
        row_total_cortes = startrow_cortes + len(df_cortes) + 1  
        
        for row_idx in [row_header_cortes, row_total_cortes]:
            for col_idx, col_name in enumerate(df_cortes.columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if 'PEND' in str(col_name).upper():
                    cell.fill = pend_fill
                else:
                    cell.fill = blue_fill
                cell.font = white_font

    # 4. Formato Tabla 2 (RF Fotos Dinámica)
    if not df_fotos_dinamica.empty:
        row_header_fotos = fila_t2 + 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_header_fotos, column=col_idx)
            if cell.value is not None:
                cell.fill = blue_fill
                cell.font = white_font
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            if cell.value is not None:
                cell.fill = blue_fill
                cell.font = white_font

    # 5. Ajustar ancho de columnas
    for col_idx in range(1, ws.max_column + 1):
        letra_columna = get_column_letter(col_idx)
        ws.column_dimensions[letra_columna].width = 15

    wb.save(ruta_salida)
    print(f"\n¡Éxito! Ambos reportes consolidados con el nuevo formato en:\n{ruta_salida}")