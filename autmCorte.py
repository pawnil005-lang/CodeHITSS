import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import os

# --- 1. CONFIGURACIÓN DE RUTAS ---
ruta_descargas = Path.home() / 'Downloads'
ruta_directorio_salida = Path(r"C:\Users\mendozapa\HITSS\Angel Jesus Zavala Ubillus - Campañas Hitss - Angel Zavala\Cortes x Día - Minimesas")
ruta_logo = ruta_directorio_salida / "LogoHits.png"
ruta_excel_salida = ruta_directorio_salida / "Corte_Autorización_Materiales.xlsx"

# Buscar archivos que contengan 'soporte_remoto' en el nombre (ignorando archivos temporales abiertos)
archivos_soporte = [f for f in ruta_descargas.glob('*soporte_remoto*.xls*') if '~$' not in f.name]

if not archivos_soporte:
    print("Error: No se encontró ningún archivo con 'soporte_remoto' en el nombre en la carpeta de Descargas.")
else:
    # Tomamos el archivo más reciente
    archivo_base = max(archivos_soporte, key=os.path.getmtime)
    print(f"Procesando archivo base: {archivo_base.name}")

    try:
        # --- 2. EXTRACCIÓN Y LIMPIEZA DE DATOS ---
        df = pd.read_excel(archivo_base)
        
        # Limpiar nombres de columnas (mayúsculas, sin espacios extra y sin tildes para evitar errores)
        df.columns = df.columns.astype(str).str.strip().str.upper()
        df.columns = df.columns.str.replace('Ó', 'O').str.replace('É', 'E').str.replace('Í', 'I').str.replace('Á', 'A')

        # --- 3. PROCESAMIENTO Y CÁLCULOS ---
        # 1. Total SOT
        total_sot = df['SOT'].notna().sum() if 'SOT' in df.columns else len(df)

        # 2. Contabilizar Atendidos y Pendientes según Estado
        if 'ESTADO' in df.columns:
            estados = df['ESTADO'].astype(str).str.strip().str.upper().str.replace('Ó', 'O')
            atendidos = estados.isin(['ATENDIDO', 'DENEGADO']).sum()
            pendientes = (estados == 'EN ATENCION').sum()
        else:
            print("Aviso: No se encontró la columna 'ESTADO'.")
            atendidos, pendientes = 0, 0

        # 3. Promedios de Tiempos
        def calcular_promedio_tiempo(nombre_columna):
            if nombre_columna not in df.columns:
                return "00:00:00"
            try:
                # Convertir a texto y luego a un formato de tiempo medible (timedelta)
                serie_tiempos = pd.to_timedelta(df[nombre_columna].astype(str).loc[df[nombre_columna].notna()], errors='coerce')
                promedio = serie_tiempos.mean()
                
                if pd.isna(promedio):
                    return "00:00:00"
                
                # Formatear el resultado a HH:MM:SS
                total_segundos = int(promedio.total_seconds())
                horas, remanente = divmod(total_segundos, 3600)
                minutos, segundos = divmod(remanente, 60)
                return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
            except Exception as e:
                return "00:00:00"

        promedio_espera = calcular_promedio_tiempo('TIEMPO DE ESPERA')
        promedio_atencion = calcular_promedio_tiempo('TIEMPO ATENCION')

        # --- 4. CREACIÓN Y FORMATO ESTÉTICO DEL NUEVO EXCEL ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Soporte Remoto"

        # 👉 ¡NUEVO!: Desactivar líneas de cuadrilla del fondo
        ws.sheet_view.showGridLines = False

        # Colores y Estilos
        color_azul_oscuro = "1F4E78"
        color_azul_claro = "5B9BD5"
        relleno_oscuro = PatternFill(start_color=color_azul_oscuro, end_color=color_azul_oscuro, fill_type="solid")
        relleno_claro = PatternFill(start_color=color_azul_claro, end_color=color_azul_claro, fill_type="solid")
        fuente_blanca_negrita = Font(color="FFFFFF", bold=True)
        alineacion_centro = Alignment(horizontal="center", vertical="center")
        
        # Borde gris un poco más sólido para que destaque al no haber cuadrilla
        borde_delgado = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
                               top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))

        # Encabezado Principal (Abarca 5 columnas)
        ws.merge_cells('A1:E3')

        ws['A1'] = f"CORTE AUTORIZACIÓN DE MATERIALES"
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=16)
        ws['A1'].alignment = alineacion_centro

        # 👉 ¡NUEVO!: Aplicar fondo y bordes a TODAS las celdas del bloque del encabezado principal
        for r in range(1, 4):
            for c in range(1, 6): # De la columna 1 a la 5
                celda_encabezado = ws.cell(row=r, column=c)
                celda_encabezado.fill = relleno_oscuro
                celda_encabezado.border = borde_delgado

        # Insertar el logo
        try:
            img_excel = Image(ruta_logo)
            img_excel.width = 60
            img_excel.height = 60
            ws.add_image(img_excel, 'A1')
        except Exception as e:
            print(f"Aviso: No se pudo cargar el logo. Verifica que 'LogoHits.png' esté en la ruta indicada. Error: {e}")

        # Cabeceras de la Tabla (5 Columnas)
        encabezados_tabla = ["CASOS", "PENDIENTES", "ATENDIDOS", "TME", "TMA"]
        for col_idx, texto in enumerate(encabezados_tabla, 1):
            celda = ws.cell(row=5, column=col_idx, value=texto)
            celda.font = fuente_blanca_negrita
            celda.alignment = alineacion_centro
            celda.fill = relleno_claro if texto == "PENDIENTES" else relleno_oscuro
            # 👉 ¡NUEVO!: Borde explícito para la cabecera
            celda.border = borde_delgado 

        # Volcar la data en la fila 6
        valores_fila = [total_sot, pendientes, atendidos, promedio_espera, promedio_atencion]
        for col_idx, val in enumerate(valores_fila, 1):
            celda = ws.cell(row=6, column=col_idx, value=val)
            celda.alignment = alineacion_centro
            # El borde ya estaba, pero ahora hace juego con el resto
            celda.border = borde_delgado 

        # Ajustar anchos de columnas para que se vea ordenado
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        # Guardar en la ruta específica
        ruta_directorio_salida.mkdir(parents=True, exist_ok=True)
        wb.save(ruta_excel_salida)
        
        print(f"¡Éxito! El cuadro estético se ha generado en:\n{ruta_excel_salida}")

    except Exception as e:
        print(f"Ocurrió un error inesperado al procesar el archivo: {e}")