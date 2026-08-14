import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image

# --- 1. CONFIGURACIÓN INICIAL ---
# 👇 AQUÍ PONES LA FECHA MANUAL DESDE LA CUAL QUIERES FILTRAR (Formato: DD/MM/AAAA) 👇
FECHA_INICIO_MANUAL = "12/08/2026"

ruta_descargas = Path.home() / 'Downloads' / 'CORTES NPS.xlsx'
ruta_directorio_salida = Path(r"C:\Users\mendozapa\HITSS\Angel Jesus Zavala Ubillus - Campañas Hitss - Angel Zavala\Cortes x Día - Minimesas")
ruta_logo = ruta_directorio_salida / "LogoHits.png"
ruta_excel_salida = ruta_directorio_salida / "Corte_TNPS_NPS.xlsx"

try:
    # --- 2. EXTRACCIÓN DE DATOS (IGNORANDO FILAS OCULTAS) ---
    print("Leyendo filas visibles del documento...")
    wb_base = openpyxl.load_workbook(ruta_descargas, data_only=True)
    ws_base = wb_base.active
    
    datos_visibles = []
    for row_idx, row in enumerate(ws_base.iter_rows(values_only=True), start=1):
        if not ws_base.row_dimensions[row_idx].hidden:
            datos_visibles.append(row)
            
    if not datos_visibles:
        print("Error: El documento no contiene datos visibles.")
    else:
        df_raw = pd.DataFrame(datos_visibles)
        encabezados = df_raw.iloc[0].astype(str).str.strip().str.upper().tolist()
        df_datos = df_raw.iloc[1:].reset_index(drop=True)

        # --- 3. AGRUPACIÓN Y LIMPIEZA DE BLOQUES ---
        bloques = []
        for i, col in enumerate(encabezados):
            if col in ['FECHA_ASIGNACIÓN', 'FECHA_ASIGNACION']:
                try:
                    # Agrupamos la fecha junto con TOTAL y GESTIONADO (índices continuos)
                    df_temp = df_datos.iloc[:, [i, i+1, i+2]].copy()
                    df_temp.columns = ['FECHA', 'TOTAL', 'GESTIONADO']
                    bloques.append(df_temp)
                except IndexError:
                    pass

        # Unir verticalmente y limpiar datos vacíos
        df_consolidado = pd.concat(bloques, ignore_index=True)
        df_consolidado = df_consolidado.dropna(subset=['FECHA'])
        df_consolidado['TOTAL'] = pd.to_numeric(df_consolidado['TOTAL'], errors='coerce').fillna(0)
        df_consolidado['GESTIONADO'] = pd.to_numeric(df_consolidado['GESTIONADO'], errors='coerce').fillna(0)
        
        # Convertir la columna de fechas a un formato de tiempo real
        df_consolidado['FECHA_DT'] = pd.to_datetime(df_consolidado['FECHA'], dayfirst=True, errors='coerce')

        # --- 4. NUEVO FILTRO: IGNORAR FECHAS ANTERIORES A LA MANUAL ---
        limite_fecha = pd.to_datetime(FECHA_INICIO_MANUAL, format="%d/%m/%Y")
        df_consolidado = df_consolidado[df_consolidado['FECHA_DT'] >= limite_fecha]

        # --- 5. CÁLCULOS FINALES ---
        # Sumar totales agrupando por fecha y calcular pendientes
        df_agrupado = df_consolidado.groupby('FECHA_DT', as_index=False)[['TOTAL', 'GESTIONADO']].sum()
        df_agrupado['PENDIENTES'] = df_agrupado['TOTAL'] - df_agrupado['GESTIONADO']
        
        # Filtrar fechas que quedaron en 0
        df_final = df_agrupado[df_agrupado['PENDIENTES'] != 0].reset_index(drop=True)

        # Formatear la fecha a "dd-mmm" (ej. 08-ago)
        meses_es = {1:'ene', 2:'feb', 3:'mar', 4:'abr', 5:'may', 6:'jun', 
                    7:'jul', 8:'ago', 9:'sep', 10:'oct', 11:'nov', 12:'dic'}
        df_final['FECHA_STR'] = df_final['FECHA_DT'].apply(
            lambda x: f"{x.day:02d}-{meses_es.get(x.month, '')}" if pd.notnull(x) else ""
        )

        # --- 6. CREACIÓN Y FORMATO DEL NUEVO EXCEL ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Corte TNPS"

        # 👉 ¡NUEVO!: Desactivar líneas de cuadrícula
        ws.sheet_view.showGridLines = False

        # Definir estilos
        color_azul_oscuro = "1F4E78"
        color_azul_claro = "5B9BD5"
        relleno_oscuro = PatternFill(start_color=color_azul_oscuro, end_color=color_azul_oscuro, fill_type="solid")
        relleno_claro = PatternFill(start_color=color_azul_claro, end_color=color_azul_claro, fill_type="solid")
        fuente_blanca_negrita = Font(color="FFFFFF", bold=True)
        alineacion_centro = Alignment(horizontal="center", vertical="center")
        
        # 👉 ¡NUEVO!: Borde gris sólido para que destaque al no haber cuadrícula
        borde_delgado = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
                               top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))

        # Encabezado Principal (Ahora se extiende hasta la columna F por las 2 nuevas columnas)
        ws.merge_cells('A1:F3')
      
        ws['A1'] = f"CORTE NPS"
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=16)
        ws['A1'].alignment = alineacion_centro

        # 👉 ¡NUEVO!: Aplicar fondo y bordes a TODAS las celdas del bloque del encabezado principal
        for r in range(1, 4):
            for c in range(1, 7): # De 1 a 6 columnas
                celda_encabezado = ws.cell(row=r, column=c)
                celda_encabezado.fill = relleno_oscuro
                celda_encabezado.border = borde_delgado

        # Insertar el logo
        try:
            img_excel = Image(ruta_logo)
            img_excel.width = 60
            img_excel.height = 60
            ws.add_image(img_excel, 'C1')
        except Exception as e:
            print(f"Aviso: No se pudo cargar el logo. Error: {e}")

        # Cabeceras de la Tabla (Se añaden % PENDIENTES y % TERMINADOS)
        encabezados_tabla = ["FECHA", "TOTAL", "PENDIENTES", "%", "TERMINADOS", "%"]
        for col_idx, texto in enumerate(encabezados_tabla, 1):
            celda = ws.cell(row=5, column=col_idx, value=texto)
            celda.font = fuente_blanca_negrita
            celda.alignment = alineacion_centro
            # Mantener celeste para todo lo de Pendientes, azul oscuro para el resto
            celda.fill = relleno_claro if "PENDIENTES" in texto else relleno_oscuro
            # 👉 ¡NUEVO!: Borde explícito para la cabecera
            celda.border = borde_delgado

        # Volcar Filas de Datos
        fila_actual = 6
        suma_total = suma_pendientes = suma_terminados = 0

        for _, row in df_final.iterrows():
            # Calcular porcentajes por fila (con control para no dividir entre 0)
            pct_pend = row['PENDIENTES'] / row['TOTAL'] if row['TOTAL'] != 0 else 0
            pct_term = row['GESTIONADO'] / row['TOTAL'] if row['TOTAL'] != 0 else 0
            
            valores_fila = [row['FECHA_STR'], row['TOTAL'], row['PENDIENTES'], pct_pend, row['GESTIONADO'], pct_term]
            
            for col_idx, val in enumerate(valores_fila, 1):
                celda = ws.cell(row=fila_actual, column=col_idx, value=val)
                celda.alignment = alineacion_centro
                celda.border = borde_delgado
                
                # Si estamos en las columnas de porcentaje (4 y 6), damos formato de porcentaje a la celda
                if col_idx in [4, 6]:
                    celda.number_format = '0%'
            
            suma_total += row['TOTAL']
            suma_pendientes += row['PENDIENTES']
            suma_terminados += row['GESTIONADO']
            fila_actual += 1

        # Calcular porcentajes totales para la última fila
        pct_total_pend = suma_pendientes / suma_total if suma_total != 0 else 0
        pct_total_term = suma_terminados / suma_total if suma_total != 0 else 0

        # Fila de TOTAL general
        celdas_total = [
            ws.cell(row=fila_actual, column=1, value="TOTAL"),
            ws.cell(row=fila_actual, column=2, value=suma_total),
            ws.cell(row=fila_actual, column=3, value=suma_pendientes),
            ws.cell(row=fila_actual, column=4, value=pct_total_pend),
            ws.cell(row=fila_actual, column=5, value=suma_terminados),
            ws.cell(row=fila_actual, column=6, value=pct_total_term)
        ]

        for i, celda in enumerate(celdas_total):
            celda.font = fuente_blanca_negrita
            celda.alignment = alineacion_centro
            celda.border = borde_delgado
            # Indices 2 y 3 corresponden a PENDIENTES y % (el que está al lado de pendientes) respectivamente
            celda.fill = relleno_claro if i in [2, 3] else relleno_oscuro
            
            # Dar formato de porcentaje a las columnas 4 y 6 de la fila total
            if i in [3, 5]:
                celda.number_format = '0%'

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18 # % Pendientes
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18 # % Terminados
        
        # Guardar archivo
        ruta_directorio_salida.mkdir(parents=True, exist_ok=True)
        wb.save(ruta_excel_salida)
        print(f"¡Éxito! El cuadro estético se ha generado en:\n{ruta_excel_salida}")

except FileNotFoundError as e:
    print(f"Error: {e}. Por favor, verifica que el archivo 'CORTES NPS.xlsx' exista en Descargas.")
except Exception as e:
    print(f"Ocurrió un error inesperado: {e}")